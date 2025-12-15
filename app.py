"""
Reconciliation Work Allocation App with Role-Based Access Control
- Admin: Full access (add/edit/delete)
- User: View and mark completion only
"""
from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, session, Response
from flask_sqlalchemy import SQLAlchemy
from werkzeug.security import generate_password_hash, check_password_hash
from functools import wraps
from datetime import datetime, date, timedelta, timezone
import os
import sys
import io

# Sri Lanka timezone (UTC+5:30)
SL_OFFSET = timedelta(hours=5, minutes=30)

def get_sl_now():
    """Get current datetime in Sri Lanka timezone"""
    return datetime.utcnow() + SL_OFFSET

def get_sl_today():
    """Get current date in Sri Lanka timezone"""
    return get_sl_now().date()

# Handle PyInstaller bundling
if getattr(sys, 'frozen', False):
    template_folder = os.path.join(sys._MEIPASS, 'templates')
    static_folder = os.path.join(sys._MEIPASS, 'static')
    app = Flask(__name__, template_folder=template_folder, static_folder=static_folder)
    # EXE: Data stored in user's home folder
    app_data = os.path.join(os.path.expanduser('~'), 'ReconciliationApp')
    os.makedirs(app_data, exist_ok=True)
    db_path = os.path.join(app_data, 'reconciliation.db')
    database_url = f'sqlite:///{db_path}'
    print(f"\n[DATA LOCATION] Database stored at: {db_path}")
else:
    app = Flask(__name__)
    # Check for Railway/Cloud DATABASE_URL first, otherwise use local SQLite
    database_url = os.environ.get('DATABASE_URL')
    if database_url:
        # Railway uses postgres://, SQLAlchemy needs postgresql+psycopg://
        if database_url.startswith('postgres://'):
            database_url = database_url.replace('postgres://', 'postgresql+psycopg://', 1)
        elif database_url.startswith('postgresql://'):
            database_url = database_url.replace('postgresql://', 'postgresql+psycopg://', 1)
        db_path = "Cloud PostgreSQL Database"
        print(f"\n[DATABASE] Using cloud PostgreSQL database")
    else:
        # Local development: use SQLite
        db_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'reconciliation.db')
        database_url = f'sqlite:///{db_path}'
        print(f"\n[DATA LOCATION] Database stored at: {db_path}")

app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'your-secret-key-change-in-production')
app.config['SQLALCHEMY_DATABASE_URI'] = database_url
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db = SQLAlchemy(app)

# Template filter to convert UTC to Sri Lanka time
@app.template_filter('local_time')
def local_time_filter(utc_dt):
    if utc_dt is None:
        return ''
    local_dt = utc_dt + SL_OFFSET
    return local_dt.strftime('%Y-%m-%d %H:%M')

@app.template_filter('local_date')
def local_date_filter(utc_dt):
    if utc_dt is None:
        return ''
    local_dt = utc_dt + SL_OFFSET
    return local_dt.strftime('%Y-%m-%d')

# ============== DATABASE MODELS ==============

class User(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    password_hash = db.Column(db.String(256), nullable=False)
    name = db.Column(db.String(100), nullable=False)
    role = db.Column(db.String(20), nullable=False, default='user')  # 'admin' or 'user'
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    
    def set_password(self, password):
        self.password_hash = generate_password_hash(password)
    
    def check_password(self, password):
        return check_password_hash(self.password_hash, password)
    
    def is_admin(self):
        return self.role == 'admin'

class TeamMember(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    email = db.Column(db.String(120), unique=True, nullable=False)
    role = db.Column(db.String(50), nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    reconciliations = db.relationship('Reconciliation', backref='assignee', lazy=True)

class Reconciliation(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(200), nullable=False)
    description = db.Column(db.Text)
    frequency = db.Column(db.String(20), nullable=False)
    priority = db.Column(db.String(20), default='Medium')
    status = db.Column(db.String(20), default='Pending')
    source_system = db.Column(db.String(100))
    target_system = db.Column(db.String(100))
    due_date = db.Column(db.Date)
    due_time = db.Column(db.String(5), default='17:00')  # Format: HH:MM (24-hour) - for Daily only
    last_completed = db.Column(db.DateTime)
    next_due = db.Column(db.Date)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    assigned_to = db.Column(db.Integer, db.ForeignKey('team_member.id'))
    completion_notes = db.Column(db.Text)
    items_reconciled = db.Column(db.Integer, default=0)
    exceptions_found = db.Column(db.Integer, default=0)
    completed_by = db.Column(db.String(100))
    overdue_notified = db.Column(db.Boolean, default=False)  # Track if overdue email was sent
    
    def is_overdue(self):
        """Check if reconciliation is overdue based on date and time (Sri Lanka timezone)"""
        if self.status == 'Completed' or not self.due_date:
            return False
        
        # Get current date and time in Sri Lanka
        now_sl = get_sl_now()
        today_sl = get_sl_today()
        
        if self.due_date < today_sl:
            return True
        elif self.due_date == today_sl:
            # For Daily recs, check due time
            if self.frequency == 'Daily' and self.due_time:
                try:
                    due_hour, due_min = map(int, self.due_time.split(':'))
                    if now_sl.hour > due_hour or (now_sl.hour == due_hour and now_sl.minute > due_min):
                        return True
                except:
                    pass
            # For Weekly/Monthly, overdue at end of day (11:59 PM)
            elif self.frequency in ['Weekly', 'Monthly']:
                if now_sl.hour >= 23 and now_sl.minute >= 59:
                    return True
        return False
    
    def is_due_today(self):
        """Check if reconciliation is due today but not yet overdue (Sri Lanka timezone)"""
        if self.status == 'Completed' or not self.due_date:
            return False
        return self.due_date == get_sl_today() and not self.is_overdue()
    
    @staticmethod
    def get_last_working_day_of_week(reference_date=None):
        """Get last working day (Friday) of the current week"""
        if reference_date is None:
            reference_date = get_sl_today()
        # Find Friday of this week
        days_until_friday = 4 - reference_date.weekday()
        if days_until_friday < 0:
            days_until_friday += 7
        friday = reference_date + timedelta(days=days_until_friday)
        return friday
    
    @staticmethod
    def get_next_last_working_day_of_week():
        """Get last working day (Friday) of next week"""
        today = get_sl_today()
        # Find Friday of next week
        days_until_friday = 4 - today.weekday()
        if days_until_friday <= 0:
            days_until_friday += 7
        next_friday = today + timedelta(days=days_until_friday)
        # If today is Friday or past Friday, go to next week's Friday
        if today.weekday() >= 4:
            next_friday += timedelta(days=7)
        return next_friday
    
    def calculate_next_due(self):
        today = get_sl_today()
        if self.frequency == 'Daily':
            # Next business day (skip weekends)
            next_day = today + timedelta(days=1)
            while next_day.weekday() >= 5:  # Saturday=5, Sunday=6
                next_day += timedelta(days=1)
            return next_day
        elif self.frequency == 'Weekly':
            # Last working day of next week (Friday)
            return self.get_next_last_working_day_of_week()
        elif self.frequency == 'Monthly':
            # 1st working day of next month
            month = today.month + 1
            year = today.year
            if month > 12:
                month = 1
                year += 1
            first_of_month = date(year, month, 1)
            # If 1st is Saturday, move to Monday (3rd)
            # If 1st is Sunday, move to Monday (2nd)
            while first_of_month.weekday() >= 5:
                first_of_month += timedelta(days=1)
            return first_of_month
        return today

class Notification(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    title = db.Column(db.String(200), nullable=False)
    message = db.Column(db.Text, nullable=False)
    type = db.Column(db.String(20), default='warning')  # info, warning, danger, success
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True)  # For specific user
    for_admins = db.Column(db.Boolean, default=False)  # Show to all admins
    for_member_id = db.Column(db.Integer, db.ForeignKey('team_member.id'), nullable=True)  # For team member
    rec_id = db.Column(db.Integer, db.ForeignKey('reconciliation.id'), nullable=True)
    is_read = db.Column(db.Boolean, default=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    
    user = db.relationship('User', backref='notifications')
    team_member = db.relationship('TeamMember', backref='notifications')
    reconciliation = db.relationship('Reconciliation', backref='notifications')

class CompletionHistory(db.Model):
    """Track all reconciliation completions for historical reporting"""
    id = db.Column(db.Integer, primary_key=True)
    reconciliation_id = db.Column(db.Integer, db.ForeignKey('reconciliation.id'), nullable=False)
    reconciliation_name = db.Column(db.String(200), nullable=False)
    frequency = db.Column(db.String(20), nullable=False)
    priority = db.Column(db.String(20))
    source_system = db.Column(db.String(100))
    target_system = db.Column(db.String(100))
    assigned_to_name = db.Column(db.String(100))
    completed_by = db.Column(db.String(100))
    due_date = db.Column(db.Date)
    completed_at = db.Column(db.DateTime, default=datetime.utcnow)
    items_reconciled = db.Column(db.Integer, default=0)
    exceptions_found = db.Column(db.Integer, default=0)
    completion_notes = db.Column(db.Text)
    was_overdue = db.Column(db.Boolean, default=False)
    days_overdue = db.Column(db.Integer, default=0)
    
    reconciliation = db.relationship('Reconciliation', backref='completion_history')

# ============== NOTIFICATION FUNCTIONS ==============

def create_overdue_notification(rec):
    """Create overdue notification for assigned member and all admins"""
    # Check if notification already exists for this rec
    existing = Notification.query.filter_by(rec_id=rec.id, type='danger', is_read=False).first()
    if existing:
        return False
    
    due_info = f"Due: {rec.due_date.strftime('%Y-%m-%d')}"
    if rec.frequency == 'Daily' and rec.due_time:
        due_info += f" at {rec.due_time}"
    
    # Create notification for admins
    admin_notif = Notification(
        title=f"OVERDUE: {rec.name}",
        message=f"{rec.name} ({rec.frequency}) is overdue! {due_info}. Assigned to: {rec.assignee.name if rec.assignee else 'Unassigned'}",
        type='danger',
        for_admins=True,
        rec_id=rec.id
    )
    db.session.add(admin_notif)
    
    # Create notification for assigned team member (linked to user if exists)
    if rec.assignee:
        member_notif = Notification(
            title=f"OVERDUE: {rec.name}",
            message=f"Your reconciliation '{rec.name}' is overdue! {due_info}. Please contact an admin to complete it.",
            type='danger',
            for_member_id=rec.assigned_to,
            rec_id=rec.id
        )
        db.session.add(member_notif)
    
    db.session.commit()
    return True

def check_and_create_overdue_notifications():
    """Check for overdue items and create notifications"""
    # Get all overdue reconciliations that haven't been notified
    overdue_recs = Reconciliation.query.filter(
        Reconciliation.status != 'Completed',
        Reconciliation.overdue_notified == False
    ).all()
    
    # Filter to actually overdue items
    overdue_recs = [r for r in overdue_recs if r.is_overdue()]
    
    for rec in overdue_recs:
        if create_overdue_notification(rec):
            rec.overdue_notified = True
    
    db.session.commit()
    return len(overdue_recs)

def get_user_notifications(user):
    """Get notifications for a specific user"""
    notifications = []
    
    # If admin, get admin notifications
    if user.is_admin():
        admin_notifs = Notification.query.filter_by(for_admins=True, is_read=False).order_by(Notification.created_at.desc()).all()
        notifications.extend(admin_notifs)
    
    # Get notifications for this user specifically
    user_notifs = Notification.query.filter_by(user_id=user.id, is_read=False).order_by(Notification.created_at.desc()).all()
    notifications.extend(user_notifs)
    
    # Get notifications for team member linked to this user (by name match)
    member = TeamMember.query.filter_by(name=user.name).first()
    if member:
        member_notifs = Notification.query.filter_by(for_member_id=member.id, is_read=False).order_by(Notification.created_at.desc()).all()
        notifications.extend(member_notifs)
    
    # Remove duplicates and sort
    seen_ids = set()
    unique_notifs = []
    for n in notifications:
        if n.id not in seen_ids:
            seen_ids.add(n.id)
            unique_notifs.append(n)
    
    return sorted(unique_notifs, key=lambda x: x.created_at, reverse=True)

# ============== AUTO-RESET FUNCTIONS ==============

def auto_reset_completed_reconciliations():
    """
    Automatically reset completed reconciliations when their next_due date arrives.
    This makes items available again for the new cycle.
    """
    today = get_sl_today()
    
    # Find all completed reconciliations where next_due <= today
    completed_recs = Reconciliation.query.filter(
        Reconciliation.status == 'Completed',
        Reconciliation.next_due <= today
    ).all()
    
    reset_count = 0
    for rec in completed_recs:
        # Reset the reconciliation for new cycle
        rec.status = 'Pending'
        rec.due_date = rec.next_due  # Set due_date to the next_due
        rec.next_due = rec.calculate_next_due()  # Calculate new next_due
        rec.items_reconciled = 0
        rec.exceptions_found = 0
        rec.completion_notes = ''
        rec.completed_by = None
        rec.overdue_notified = False
        reset_count += 1
    
    if reset_count > 0:
        db.session.commit()
        print(f"[AUTO-RESET] Reset {reset_count} reconciliation(s) for new cycle")
    
    return reset_count

# ============== ACCESS CONTROL ==============

def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            flash('Please log in to access this page.', 'error')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            flash('Please log in to access this page.', 'error')
            return redirect(url_for('login'))
        user = User.query.get(session['user_id'])
        if not user or not user.is_admin():
            flash('You do not have permission to perform this action. Admin access required.', 'error')
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)
    return decorated_function

def get_current_user():
    if 'user_id' in session:
        return User.query.get(session['user_id'])
    return None

@app.context_processor
def inject_user():
    return dict(current_user=get_current_user())

# ============== AUTH ROUTES ==============

@app.route('/login', methods=['GET', 'POST'])
def login():
    if 'user_id' in session:
        return redirect(url_for('dashboard'))
    
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        user = User.query.filter_by(username=username).first()
        
        if user and user.check_password(password):
            session['user_id'] = user.id
            flash(f'Welcome back, {user.name}!', 'success')
            return redirect(url_for('dashboard'))
        else:
            flash('Invalid username or password.', 'error')
    
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.pop('user_id', None)
    flash('You have been logged out.', 'success')
    return redirect(url_for('login'))

@app.route('/change-password', methods=['GET', 'POST'])
@login_required
def change_password():
    if request.method == 'POST':
        current_user = get_current_user()
        old_password = request.form['old_password']
        new_password = request.form['new_password']
        confirm_password = request.form['confirm_password']
        
        if not current_user.check_password(old_password):
            flash('Current password is incorrect.', 'error')
        elif new_password != confirm_password:
            flash('New passwords do not match.', 'error')
        elif len(new_password) < 4:
            flash('Password must be at least 4 characters.', 'error')
        else:
            current_user.set_password(new_password)
            db.session.commit()
            flash('Password changed successfully!', 'success')
            return redirect(url_for('dashboard'))
    
    return render_template('change_password.html')

# ============== USER MANAGEMENT (Admin Only) ==============

@app.route('/users')
@admin_required
def list_users():
    users = User.query.all()
    return render_template('users.html', users=users)

@app.route('/users/add', methods=['GET', 'POST'])
@admin_required
def add_user():
    if request.method == 'POST':
        username = request.form['username']
        name = request.form['name']
        password = request.form['password']
        role = request.form['role']
        
        if User.query.filter_by(username=username).first():
            flash('Username already exists!', 'error')
            return redirect(url_for('add_user'))
        
        user = User(username=username, name=name, role=role)
        user.set_password(password)
        db.session.add(user)
        db.session.commit()
        flash(f'User "{name}" created successfully!', 'success')
        return redirect(url_for('list_users'))
    
    return render_template('add_user.html')

@app.route('/users/edit/<int:id>', methods=['GET', 'POST'])
@admin_required
def edit_user(id):
    user = User.query.get_or_404(id)
    if request.method == 'POST':
        user.name = request.form['name']
        user.role = request.form['role']
        if request.form.get('password'):
            user.set_password(request.form['password'])
        db.session.commit()
        flash(f'User "{user.name}" updated successfully!', 'success')
        return redirect(url_for('list_users'))
    
    return render_template('edit_user.html', user=user)

@app.route('/users/delete/<int:id>')
@admin_required
def delete_user(id):
    user = User.query.get_or_404(id)
    if user.id == session['user_id']:
        flash('You cannot delete your own account!', 'error')
        return redirect(url_for('list_users'))
    db.session.delete(user)
    db.session.commit()
    flash(f'User "{user.name}" deleted successfully!', 'success')
    return redirect(url_for('list_users'))

# ============== DASHBOARD ==============

@app.route('/')
@login_required
def dashboard():
    # Auto-reset completed reconciliations that are due
    auto_reset_completed_reconciliations()
    
    today = get_sl_today()
    
    total_members = TeamMember.query.count()
    total_recs = Reconciliation.query.count()
    
    daily_recs = Reconciliation.query.filter_by(frequency='Daily').count()
    weekly_recs = Reconciliation.query.filter_by(frequency='Weekly').count()
    monthly_recs = Reconciliation.query.filter_by(frequency='Monthly').count()
    
    pending_recs = Reconciliation.query.filter_by(status='Pending').count()
    in_progress_recs = Reconciliation.query.filter_by(status='In Progress').count()
    completed_recs = Reconciliation.query.filter_by(status='Completed').count()
    on_hold_recs = Reconciliation.query.filter_by(status='On Hold').count()
    
    due_today = Reconciliation.query.filter(
        Reconciliation.due_date == today,
        Reconciliation.status != 'Completed'
    ).all()
    
    # Filter due_today to exclude items that are actually overdue (for daily items with time)
    due_today = [r for r in due_today if not r.is_overdue()]
    
    # Get all potential overdue items (due_date < today OR due_date == today but past due_time)
    potential_overdue = Reconciliation.query.filter(
        Reconciliation.due_date <= today,
        Reconciliation.status != 'Completed'
    ).all()
    
    # Filter to only actually overdue items using the is_overdue method
    overdue = [r for r in potential_overdue if r.is_overdue()]
    
    # Create notifications for new overdue items
    check_and_create_overdue_notifications()
    
    members = TeamMember.query.all()
    workload = []
    for member in members:
        active_recs = Reconciliation.query.filter(
            Reconciliation.assigned_to == member.id,
            Reconciliation.status.in_(['Pending', 'In Progress'])
        ).count()
        workload.append({'member': member, 'active_recs': active_recs})
    
    recent_completed = Reconciliation.query.filter_by(status='Completed').order_by(
        Reconciliation.last_completed.desc()
    ).limit(5).all()
    
    return render_template('dashboard.html',
                         total_members=total_members, total_recs=total_recs,
                         daily_recs=daily_recs, weekly_recs=weekly_recs, monthly_recs=monthly_recs,
                         pending_recs=pending_recs, in_progress_recs=in_progress_recs,
                         completed_recs=completed_recs, on_hold_recs=on_hold_recs,
                         due_today=due_today, overdue=overdue,
                         workload=workload, recent_completed=recent_completed, today=today)

# ============== TEAM MEMBER ROUTES ==============

@app.route('/members')
@login_required
def list_members():
    members = TeamMember.query.all()
    return render_template('members.html', members=members)

@app.route('/members/add', methods=['GET', 'POST'])
@admin_required
def add_member():
    if request.method == 'POST':
        name = request.form['name']
        email = request.form['email']
        role = request.form['role']
        
        if TeamMember.query.filter_by(email=email).first():
            flash('Email already exists!', 'error')
            return redirect(url_for('add_member'))
        
        member = TeamMember(name=name, email=email, role=role)
        db.session.add(member)
        db.session.commit()
        flash(f'Team member "{name}" added successfully!', 'success')
        return redirect(url_for('list_members'))
    
    return render_template('add_member.html')

@app.route('/members/edit/<int:id>', methods=['GET', 'POST'])
@admin_required
def edit_member(id):
    member = TeamMember.query.get_or_404(id)
    if request.method == 'POST':
        member.name = request.form['name']
        member.email = request.form['email']
        member.role = request.form['role']
        db.session.commit()
        flash(f'Team member "{member.name}" updated successfully!', 'success')
        return redirect(url_for('list_members'))
    
    return render_template('edit_member.html', member=member)

@app.route('/members/delete/<int:id>')
@admin_required
def delete_member(id):
    member = TeamMember.query.get_or_404(id)
    Reconciliation.query.filter_by(assigned_to=id).update({'assigned_to': None})
    db.session.delete(member)
    db.session.commit()
    flash(f'Team member "{member.name}" deleted successfully!', 'success')
    return redirect(url_for('list_members'))

# ============== RECONCILIATION ROUTES ==============

@app.route('/reconciliations')
@login_required
def list_reconciliations():
    # Auto-reset completed reconciliations that are due
    auto_reset_completed_reconciliations()
    
    status_filter = request.args.get('status', '')
    frequency_filter = request.args.get('frequency', '')
    priority_filter = request.args.get('priority', '')
    member_filter = request.args.get('member', '')
    
    query = Reconciliation.query
    
    if status_filter:
        query = query.filter_by(status=status_filter)
    if frequency_filter:
        query = query.filter_by(frequency=frequency_filter)
    if priority_filter:
        query = query.filter_by(priority=priority_filter)
    if member_filter:
        query = query.filter_by(assigned_to=int(member_filter))
    
    recs = query.order_by(Reconciliation.due_date.asc()).all()
    members = TeamMember.query.all()
    today = get_sl_today()
    
    return render_template('reconciliations.html', recs=recs, members=members,
                         status_filter=status_filter, frequency_filter=frequency_filter,
                         priority_filter=priority_filter, member_filter=member_filter, today=today)

@app.route('/reconciliations/add', methods=['GET', 'POST'])
@admin_required
def add_reconciliation():
    if request.method == 'POST':
        name = request.form['name']
        description = request.form.get('description', '')
        frequency = request.form['frequency']
        priority = request.form['priority']
        source_system = request.form.get('source_system', '')
        target_system = request.form.get('target_system', '')
        assigned_to = request.form.get('assigned_to')
        
        # Handle due date/time based on frequency
        due_date = None
        due_time = '17:00'  # Default
        
        if frequency == 'Daily':
            # Auto-set due date to next working day, user provides due time
            today = get_sl_today()
            due_date = today
            if today.weekday() >= 5:  # Weekend
                due_date = today + timedelta(days=(7 - today.weekday()))
            due_time = request.form.get('due_time', '17:00')
            
        elif frequency == 'Weekly':
            # Auto-set due date to last working day of week (Friday)
            due_date = Reconciliation.get_last_working_day_of_week()
            due_time = None  # Not used for weekly
            
        elif frequency == 'Monthly':
            # User provides due date
            due_date_str = request.form.get('due_date')
            due_date = datetime.strptime(due_date_str, '%Y-%m-%d').date() if due_date_str else None
            due_time = None  # Not used for monthly
        
        rec = Reconciliation(
            name=name, description=description, frequency=frequency,
            priority=priority, source_system=source_system, target_system=target_system,
            assigned_to=int(assigned_to) if assigned_to else None, 
            due_date=due_date, due_time=due_time
        )
        db.session.add(rec)
        db.session.commit()
        flash(f'Reconciliation "{name}" created successfully!', 'success')
        return redirect(url_for('list_reconciliations'))
    
    members = TeamMember.query.all()
    today = get_sl_today()
    next_friday = Reconciliation.get_last_working_day_of_week()
    return render_template('add_reconciliation.html', members=members, today=today, next_friday=next_friday)

@app.route('/reconciliations/edit/<int:id>', methods=['GET', 'POST'])
@admin_required
def edit_reconciliation(id):
    rec = Reconciliation.query.get_or_404(id)
    if request.method == 'POST':
        rec.name = request.form['name']
        rec.description = request.form.get('description', '')
        old_frequency = rec.frequency
        rec.frequency = request.form['frequency']
        rec.priority = request.form['priority']
        rec.status = request.form['status']
        rec.source_system = request.form.get('source_system', '')
        rec.target_system = request.form.get('target_system', '')
        assigned_to = request.form.get('assigned_to')
        rec.assigned_to = int(assigned_to) if assigned_to else None
        
        # Handle due date/time based on frequency
        if rec.frequency == 'Daily':
            rec.due_time = request.form.get('due_time', '17:00')
            # Only recalculate due date if frequency changed
            if old_frequency != 'Daily':
                today = get_sl_today()
                rec.due_date = today
                if today.weekday() >= 5:
                    rec.due_date = today + timedelta(days=(7 - today.weekday()))
        elif rec.frequency == 'Weekly':
            rec.due_time = None
            # Only recalculate due date if frequency changed
            if old_frequency != 'Weekly':
                rec.due_date = Reconciliation.get_last_working_day_of_week()
        elif rec.frequency == 'Monthly':
            rec.due_time = None
            due_date_str = request.form.get('due_date')
            rec.due_date = datetime.strptime(due_date_str, '%Y-%m-%d').date() if due_date_str else None
        
        db.session.commit()
        flash(f'Reconciliation "{rec.name}" updated successfully!', 'success')
        return redirect(url_for('list_reconciliations'))
    
    members = TeamMember.query.all()
    today = get_sl_today()
    next_friday = Reconciliation.get_last_working_day_of_week()
    return render_template('edit_reconciliation.html', rec=rec, members=members, today=today, next_friday=next_friday)

@app.route('/reconciliations/delete/<int:id>')
@admin_required
def delete_reconciliation(id):
    rec = Reconciliation.query.get_or_404(id)
    db.session.delete(rec)
    db.session.commit()
    flash(f'Reconciliation "{rec.name}" deleted successfully!', 'success')
    return redirect(url_for('list_reconciliations'))

@app.route('/reconciliations/view/<int:id>')
@login_required
def view_reconciliation(id):
    rec = Reconciliation.query.get_or_404(id)
    return render_template('view_reconciliation.html', rec=rec)

@app.route('/reconciliations/complete/<int:id>', methods=['GET', 'POST'])
@login_required
def complete_reconciliation(id):
    rec = Reconciliation.query.get_or_404(id)
    current_user = get_current_user()
    
    # If overdue, only admin can complete
    if rec.is_overdue() and not current_user.is_admin():
        flash('This reconciliation is overdue. Only an administrator can mark it as complete.', 'error')
        return redirect(url_for('list_reconciliations'))
    
    if request.method == 'POST':
        # Calculate if it was overdue and by how many days
        was_overdue = rec.is_overdue()
        days_overdue = 0
        if rec.due_date and rec.due_date < get_sl_today():
            days_overdue = (get_sl_today() - rec.due_date).days
        
        # Save completion history before updating
        history = CompletionHistory(
            reconciliation_id=rec.id,
            reconciliation_name=rec.name,
            frequency=rec.frequency,
            priority=rec.priority,
            source_system=rec.source_system,
            target_system=rec.target_system,
            assigned_to_name=rec.assignee.name if rec.assignee else 'Unassigned',
            completed_by=current_user.name if current_user else 'Unknown',
            due_date=rec.due_date,
            completed_at=datetime.utcnow(),
            items_reconciled=int(request.form.get('items_reconciled', 0)),
            exceptions_found=int(request.form.get('exceptions_found', 0)),
            completion_notes=request.form.get('completion_notes', ''),
            was_overdue=was_overdue,
            days_overdue=days_overdue
        )
        db.session.add(history)
        
        rec.status = 'Completed'
        rec.last_completed = datetime.utcnow()
        rec.items_reconciled = int(request.form.get('items_reconciled', 0))
        rec.exceptions_found = int(request.form.get('exceptions_found', 0))
        rec.completion_notes = request.form.get('completion_notes', '')
        rec.completed_by = current_user.name if current_user else 'Unknown'
        rec.next_due = rec.calculate_next_due()
        rec.overdue_notified = False  # Reset for next cycle
        db.session.commit()
        flash(f'Reconciliation "{rec.name}" marked as complete!', 'success')
        return redirect(url_for('list_reconciliations'))
    
    return render_template('complete_reconciliation.html', rec=rec)

@app.route('/reconciliations/start/<int:id>')
@login_required
def start_reconciliation(id):
    rec = Reconciliation.query.get_or_404(id)
    rec.status = 'In Progress'
    db.session.commit()
    flash(f'Started working on "{rec.name}"!', 'success')
    return redirect(request.referrer or url_for('list_reconciliations'))

@app.route('/reconciliations/status/<int:id>/<status>')
@login_required
def update_rec_status(id, status):
    rec = Reconciliation.query.get_or_404(id)
    rec.status = status
    db.session.commit()
    flash(f'Reconciliation "{rec.name}" status updated to {status}!', 'success')
    return redirect(url_for('list_reconciliations'))

@app.route('/reconciliations/reset/<int:id>')
@admin_required
def reset_reconciliation(id):
    rec = Reconciliation.query.get_or_404(id)
    rec.status = 'Pending'
    rec.due_date = rec.next_due or rec.calculate_next_due()
    rec.items_reconciled = 0
    rec.exceptions_found = 0
    rec.completion_notes = ''
    rec.completed_by = None
    rec.overdue_notified = False  # Reset notification flag
    db.session.commit()
    flash(f'Reconciliation "{rec.name}" reset for next cycle!', 'success')
    return redirect(url_for('list_reconciliations'))

@app.route('/reconciliations/notify/<int:id>')
@admin_required
def send_notification(id):
    """Manually create overdue notification for a specific reconciliation"""
    rec = Reconciliation.query.get_or_404(id)
    
    if not rec.is_overdue():
        flash('This reconciliation is not overdue.', 'error')
        return redirect(url_for('list_reconciliations'))
    
    if create_overdue_notification(rec):
        rec.overdue_notified = True
        db.session.commit()
        flash(f'Overdue notification created for "{rec.name}"!', 'success')
    else:
        flash('Notification already exists for this item.', 'info')
    
    return redirect(url_for('list_reconciliations'))

@app.route('/api/check-overdue')
def api_check_overdue():
    """API endpoint to check and create notifications for overdue items"""
    # Auto-reset completed reconciliations that are due
    reset_count = auto_reset_completed_reconciliations()
    notif_count = check_and_create_overdue_notifications()
    return jsonify({
        'status': 'ok', 
        'message': f'Reset {reset_count} items, created {notif_count} notifications', 
        'reset_count': reset_count,
        'notification_count': notif_count
    })

@app.route('/api/notifications')
@login_required
def api_get_notifications():
    """Get notifications for current user"""
    user = get_current_user()
    if not user:
        return jsonify({'notifications': [], 'count': 0})
    
    # Auto-reset completed reconciliations and check for overdue
    auto_reset_completed_reconciliations()
    check_and_create_overdue_notifications()
    
    notifications = get_user_notifications(user)
    
    notif_list = []
    for n in notifications[:10]:  # Limit to 10 most recent
        notif_list.append({
            'id': n.id,
            'title': n.title,
            'message': n.message,
            'type': n.type,
            'rec_id': n.rec_id,
            'created_at': (n.created_at + SL_OFFSET).strftime('%Y-%m-%d %H:%M'),
            'is_read': n.is_read
        })
    
    return jsonify({
        'notifications': notif_list,
        'count': len(notifications)
    })

@app.route('/api/notifications/read/<int:id>', methods=['POST'])
@login_required
def api_mark_notification_read(id):
    """Mark a notification as read"""
    notif = Notification.query.get_or_404(id)
    notif.is_read = True
    db.session.commit()
    return jsonify({'status': 'ok'})

@app.route('/api/notifications/read-all', methods=['POST'])
@login_required
def api_mark_all_read():
    """Mark all notifications as read for current user"""
    user = get_current_user()
    if not user:
        return jsonify({'status': 'error'})
    
    notifications = get_user_notifications(user)
    for n in notifications:
        n.is_read = True
    db.session.commit()
    return jsonify({'status': 'ok'})

@app.route('/notifications')
@login_required
def view_notifications():
    """View all notifications page"""
    user = get_current_user()
    notifications = get_user_notifications(user)
    return render_template('notifications.html', notifications=notifications)

# ============== FREQUENCY VIEWS ==============

@app.route('/daily')
@login_required
def daily_view():
    today = get_sl_today()
    daily_recs = Reconciliation.query.filter_by(frequency='Daily').order_by(Reconciliation.status.asc()).all()
    members = TeamMember.query.all()
    return render_template('frequency_view.html', recs=daily_recs, members=members, today=today,
                         frequency='Daily', title='Daily Reconciliations')

@app.route('/weekly')
@login_required
def weekly_view():
    today = get_sl_today()
    weekly_recs = Reconciliation.query.filter_by(frequency='Weekly').order_by(Reconciliation.due_date.asc()).all()
    members = TeamMember.query.all()
    return render_template('frequency_view.html', recs=weekly_recs, members=members, today=today,
                         frequency='Weekly', title='Weekly Reconciliations')

@app.route('/monthly')
@login_required
def monthly_view():
    today = get_sl_today()
    monthly_recs = Reconciliation.query.filter_by(frequency='Monthly').order_by(Reconciliation.due_date.asc()).all()
    members = TeamMember.query.all()
    return render_template('frequency_view.html', recs=monthly_recs, members=members, today=today,
                         frequency='Monthly', title='Monthly Reconciliations')

# ============== COMPLETION HISTORY & REPORTS (Admin Only) ==============

@app.route('/history')
@admin_required
def completion_history():
    """View completion history with filters"""
    # Get filter parameters
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    frequency_filter = request.args.get('frequency', '')
    member_filter = request.args.get('member', '')
    overdue_only = request.args.get('overdue_only', '') == 'true'
    
    query = CompletionHistory.query
    
    if date_from:
        try:
            from_date = datetime.strptime(date_from, '%Y-%m-%d')
            query = query.filter(CompletionHistory.completed_at >= from_date)
        except ValueError:
            pass
    
    if date_to:
        try:
            to_date = datetime.strptime(date_to, '%Y-%m-%d') + timedelta(days=1)
            query = query.filter(CompletionHistory.completed_at < to_date)
        except ValueError:
            pass
    
    if frequency_filter:
        query = query.filter(CompletionHistory.frequency == frequency_filter)
    
    if member_filter:
        query = query.filter(CompletionHistory.assigned_to_name == member_filter)
    
    if overdue_only:
        query = query.filter(CompletionHistory.was_overdue == True)
    
    history = query.order_by(CompletionHistory.completed_at.desc()).all()
    
    # Get unique member names for filter dropdown
    members = db.session.query(CompletionHistory.assigned_to_name).distinct().all()
    member_names = [m[0] for m in members if m[0]]
    
    # Calculate summary stats
    total_completions = len(history)
    total_overdue = sum(1 for h in history if h.was_overdue)
    total_items = sum(h.items_reconciled or 0 for h in history)
    total_exceptions = sum(h.exceptions_found or 0 for h in history)
    
    return render_template('history.html', 
                         history=history,
                         member_names=member_names,
                         date_from=date_from,
                         date_to=date_to,
                         frequency_filter=frequency_filter,
                         member_filter=member_filter,
                         overdue_only=overdue_only,
                         total_completions=total_completions,
                         total_overdue=total_overdue,
                         total_items=total_items,
                         total_exceptions=total_exceptions)

@app.route('/history/export')
@admin_required
def export_history():
    """Export completion history to Excel"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        flash('Excel export requires openpyxl. Please install it: pip install openpyxl', 'error')
        return redirect(url_for('completion_history'))
    
    # Get filter parameters (same as history view)
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    frequency_filter = request.args.get('frequency', '')
    member_filter = request.args.get('member', '')
    overdue_only = request.args.get('overdue_only', '') == 'true'
    
    query = CompletionHistory.query
    
    if date_from:
        try:
            from_date = datetime.strptime(date_from, '%Y-%m-%d')
            query = query.filter(CompletionHistory.completed_at >= from_date)
        except ValueError:
            pass
    
    if date_to:
        try:
            to_date = datetime.strptime(date_to, '%Y-%m-%d') + timedelta(days=1)
            query = query.filter(CompletionHistory.completed_at < to_date)
        except ValueError:
            pass
    
    if frequency_filter:
        query = query.filter(CompletionHistory.frequency == frequency_filter)
    
    if member_filter:
        query = query.filter(CompletionHistory.assigned_to_name == member_filter)
    
    if overdue_only:
        query = query.filter(CompletionHistory.was_overdue == True)
    
    history = query.order_by(CompletionHistory.completed_at.desc()).all()
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Completion History"
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1e293b", end_color="1e293b", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    overdue_fill = PatternFill(start_color="fef2f2", end_color="fef2f2", fill_type="solid")
    
    # Headers
    headers = [
        "ID", "Reconciliation Name", "Frequency", "Priority", 
        "Source System", "Target System", "Assigned To", "Completed By",
        "Due Date", "Completed At", "Items Reconciled", "Exceptions Found",
        "Was Overdue", "Days Overdue", "Notes"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Data rows
    for row, h in enumerate(history, 2):
        data = [
            h.id,
            h.reconciliation_name,
            h.frequency,
            h.priority,
            h.source_system or '',
            h.target_system or '',
            h.assigned_to_name or '',
            h.completed_by or '',
            h.due_date.strftime('%Y-%m-%d') if h.due_date else '',
            (h.completed_at + SL_OFFSET).strftime('%Y-%m-%d %H:%M') if h.completed_at else '',
            h.items_reconciled or 0,
            h.exceptions_found or 0,
            'Yes' if h.was_overdue else 'No',
            h.days_overdue or 0,
            h.completion_notes or ''
        ]
        
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")
            if h.was_overdue:
                cell.fill = overdue_fill
    
    # Adjust column widths
    column_widths = [6, 30, 10, 10, 15, 15, 15, 15, 12, 18, 15, 15, 12, 12, 40]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    
    # Freeze header row
    ws.freeze_panes = 'A2'
    
    # Add summary sheet
    ws_summary = wb.create_sheet(title="Summary")
    ws_summary['A1'] = "Completion History Summary"
    ws_summary['A1'].font = Font(bold=True, size=14)
    
    ws_summary['A3'] = "Total Completions:"
    ws_summary['B3'] = len(history)
    ws_summary['A4'] = "Total Overdue:"
    ws_summary['B4'] = sum(1 for h in history if h.was_overdue)
    ws_summary['A5'] = "Total Items Reconciled:"
    ws_summary['B5'] = sum(h.items_reconciled or 0 for h in history)
    ws_summary['A6'] = "Total Exceptions Found:"
    ws_summary['B6'] = sum(h.exceptions_found or 0 for h in history)
    
    if date_from or date_to:
        ws_summary['A8'] = "Date Range:"
        ws_summary['B8'] = f"{date_from or 'Beginning'} to {date_to or 'Present'}"
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Generate filename
    filename = f"reconciliation_history_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return Response(
        output.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@app.route('/api/notifications/dismiss/<int:id>', methods=['POST'])
@login_required
def api_dismiss_notification(id):
    """Dismiss/delete a notification completely"""
    notif = Notification.query.get_or_404(id)
    db.session.delete(notif)
    db.session.commit()
    return jsonify({'status': 'ok', 'message': 'Notification dismissed'})

# ============== SETTINGS ==============

@app.route('/settings')
@login_required
def settings():
    return render_template('settings.html', db_path=db_path)

# ============== INITIALIZE ==============

def init_db():
    with app.app_context():
        db.create_all()
        if User.query.count() == 0:
            admin = User(username='admin', name='Administrator', role='admin')
            admin.set_password('admin123')
            db.session.add(admin)
            db.session.commit()
            print("\n" + "="*50)
            print("DEFAULT ADMIN ACCOUNT CREATED")
            print("="*50)
            print("Username: admin")
            print("Password: admin123")
            print("PLEASE CHANGE THIS PASSWORD AFTER FIRST LOGIN!")
            print("="*50 + "\n")

# Initialize database when app is imported (for gunicorn/production)
init_db()

if __name__ == '__main__':
    init_db()
    print("\n" + "="*50)
    print("Reconciliation App is running!")
    print("Open your browser and go to: http://127.0.0.1:5000")
    print("="*50)
    print(f"\nDATA STORED AT: {db_path}\n")
    app.run(debug=False, host='127.0.0.1', port=5000)
